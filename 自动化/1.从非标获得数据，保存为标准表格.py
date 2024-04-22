import pandas as pdimport numpy as npimport openpyxlimport picklefrom openpyxl import load_workbook# 读取Excel文件df = pd.read_excel('./非标准表格.xlsx', engine='openpyxl')#print(df)# 假设“期末余额”是列的标题，我们找到这个标题所在的所有列balance_columns = [column_name for column_name in df.columns if '期末余额' in column_name]if not balance_columns:    print("未找到包含'期末余额'的列标题。")    exit()# 创建一个空列表来存储结果result = []# 遍历所有包含“期末余额”的列for column_name in balance_columns:    # 遍历该列中所有的单元格    for index, value in enumerate(df[column_name]):        try:            # 尝试将值转换为数字，如果成功，说明是数字            num_value = pd.to_numeric(value, downcast='float')            if not pd.isna(num_value) and num_value != 0:  # 检查转换后的值是否为非数字(NaN)且不为0                # 获取对应行的左边第二列的值                left_second_column_value = df.iloc[index, df.columns.get_loc(column_name) - 2]                result.append((left_second_column_value, num_value))        except ValueError:            # 如果转换失败，说明不是数字，跳过            continue#result_dict=dict(result)#print(result)# 寻找货币资金的值currency_funds_value = Nonefor item in result:    if item[0] == '货币资金':        currency_funds_value = item[1]        break# 检查货币资金的值是否找到if currency_funds_value is not None:    cell_mapping = {        '货币资金': 'C6',        '应收账款': 'C9',        '预付款项': 'C10',        '预付账款': 'C10',        '其他应收款': 'C13',        '存货': 'C14',        '固定资产':'C24',        '无形资产':'C31',        '长期待摊费用':'C33',        '短期借款':'G6',        '应付账款':'G8',        '预收款项':'G9',        '预收账款':'G9',        '应付职工薪酬':'G10',        '应交税费':'G11',        '其他应付款':'G13',        '实收资本（或股本）':'G31',        '资本公积':'G32',        '未分配利润':'G34',        # ... 其他项目    }    try:        existing_file_path = './标准表格.xlsx'  # 替换为实际的文件路径        new_file_path = './标准表格_new.xlsx'  # 替换为你想保存的新文件路径        # 使用openpyxl加载工作簿        wb = load_workbook(existing_file_path)        ws = wb.active  # 或者使用 wb.get_sheet_by_name('Sheet1') 如果你知道工作表的名字        # 遍历每个财务项目，并将值写入到对应的单元格        for item in result:            project_name, value = item            if project_name in cell_mapping:                cell_address = cell_mapping[project_name]                ws[cell_address] = value        # 另存为新文件        wb.save(new_file_path)        print(f"数据已成功写入到新的文件'{new_file_path}'")    except FileNotFoundError:        print("指定的现有文件路径不存在，请检查文件路径。")